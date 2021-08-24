
import serial,time
from netmiko import Netmiko
from openpyxl import load_workbook

# R1 Conf
from main import console

work_book=load_workbook(filename="Routers_data.xlsx")
sheet=work_book["Sheet1"]
host1=str(sheet["F2"].value).split("/")[0]
username1=sheet["B2"].value

en_password1=sheet["D2"].value
Interface1=sheet["E4"].value
ipAdress1=str(sheet["F4"].value).split("/")[0]
password1=sheet["C2"].value
RouterInit=["enable","configure terminal","hostname R1","enable secret " + str(en_password1),"ip domain-name local","crypto key generate rsa 2048","ip ssh version 2","username " + str(username1) + "password " + str(password1),"line vty 0 4","transport input telnet ssh","login local"]
if console.isOpen():

    print("Connected successfully")
    for i in range(len(RouterInit)):

        console.write(b"i \n")

        time.sleep(1)
        console.write(b"ip address  169.254.49.202  255.255.255.0 \n")
        time.sleep(1)
        console.write(b"no shutdown\n")
        time.sleep(1)
        console.write(b'ip route 10.1.1.0 255.255.255.0 192.168.11.2\n')
        time.sleep(1)

        numberofBytes = console.inWaiting()
        data = console.read(numberofBytes)
        print(data.decode())

    else:
        print("You can't connect")

mycon=Netmiko(host="169.254.49.202", username=str(username1), password=str(password1), device_type="cisco_ios", secret= str(en_password1))
mycon.enable()
commands=["interface ethernet 0/" + str(Interface1) ,"ip address  "+ipAdress1  + " 255.255.255.0","no shutdown"]
print(commands)
output=mycon.send_config_set(commands)
print(output)
mycon.disconnect()

# R2 Conf
host2=str(sheet["F3"].value).split("/")[0]
username2=sheet["B3"].value
en_password2=sheet["D3"].value
Interface2=sheet["E5"].value
ipAdress2=str(sheet["F5"].value).split("/")[0]
password2=sheet["C3"].value

RouterInit2=["enable","configure terminal","hostname R2","enable secret " + str(en_password2),"ip domain-name local","crypto key generate rsa 2048","ip ssh version 2","username " + str(username2) + "password " + str(password2),"line vty 0 4","transport input telnet ssh","login local"]
with serial.Serial(port="COM4") as console:

  if console.isOpen():

    print("Connected successfully")
    for i in range(len(RouterInit2)):

       console.write(b"i \n")

       time.sleep(1)
       console.write(b"ip address  169.254.49.203  255.255.255.0 \n")
       time.sleep(1)
       console.write(b"no shutdown\n")
       time.sleep(1)
       console.write(b'ip route 192.168.11.1 255.255.255.0 192.168.11.2\n')
       time.sleep(1)

       numberofBytes = console.inWaiting()
       data = console.read(numberofBytes)
       print(data.decode())

    else:
        print("You can't connect")

mycon=Netmiko(host="169.254.49.203", username=str(username2), password=str(password2), device_type="cisco_ios", secret= str(en_password2))
mycon.enable()
commands=["interface ethernet 0/" + str(Interface2) ,"ip address  "+ipAdress2 + " 255.255.255.0","no shutdown"]
print(commands)
# output=mycon.send_config_set(commands)
# print(output)
mycon.disconnect()

# SW1

work_book1=load_workbook(filename="Switches_SSH-data.xlsx")
sheet=work_book1["Sheet1"]
hostS1=str(sheet["F2"].value).split("/")[0]
usernameS1=sheet["B2"].value
en_passwordS1=sheet["D2"].value
passwordS1=sheet["C2"].value
work_book2=load_workbook(filename="switch 1 Ports.xlsx")
sheet1=work_book2["Sheet1"]

SWInit=["enable","configure terminal","hostname SW1","crypto key generate rsa","login local","password " + str(passwordS1),"exit","username " + str(usernameS1) + "password " + str(passwordS1)]
with serial.Serial(port="COM4") as console:

   if console.isOpen():
        print("Connected successfully")
        for i in range(len(SWInit)):
         console.write(b"i \n")

         time.sleep(1)
         console.write(b"ip address  169.254.49.202  255.255.255.0 \n")
         time.sleep(1)
         console.write(b"no shutdown\n")
         time.sleep(1)

#          numberofBytes = console.inWaiting()
#          data = console.read(numberofBytes)
#          print(data.decode())
#         print(lis)
#     else:
#         print("You can't connect")

# mycon=Netmiko(host=hostS1, username=str(usernameS1), password=str(passwordS1), device_type="cisco_ios", secret= str(en_passwordS1))
# mycon.enable()

for x in range(2,11):

    port_number = sheet1["A" + str(x)].value
    port_type = sheet1["B" + str(x)].value
    if(port_type == "Access"):


        commAccess = ["vlan " + str(sheet1["C" + str(x)].value),"exit", "interface ethernet 0/" + str(port_number), "switchport mode access" ,"switchport access vlan " + str(sheet1["C" + str(x)].value)]
        # output = mycon.send_config_set(commAccess)
        print(commAccess)
    else:
        commTrunk = ["interface ethernet 0/" + str(port_number), "switchport trunk encapsulation dot1q" , "switchport mode trunk","switchport trunk allowed Vlan " + str(sheet1["C" + str(x)].value)]
        # output = mycon.send_config_set(commTrunk)
        print(commTrunk)


    mycon.disconnect()


#
#
#
import serial, time

lis=[]
comm=["en","terminal length 0","show ip int br","int vlan 1"]
with serial.Serial(port="COM4") as console:

    if console.isOpen():
        print("Connected successfully")
        for i in range(len(comm)):
         console.write(b"i \n")

         time.sleep(1)
         console.write(b"ip address  " + hostS1 + " 255.255.255.0 \n")
         time.sleep(1)
         console.write(b"no shutdown\n")
         time.sleep(1)
         console.write(b"ip default-gateway" + host1 + "\n")
         time.sleep(1)
         numberofBytes = console.inWaiting()
         data = console.read(numberofBytes)
         print(data.decode())
        print(lis)
    else:
        print("You can't connect")

# #
# #
# # # SW2
work_book1=load_workbook(filename="Switches_SSH-data.xlsx")
sheet=work_book1["Sheet1"]
hostS2=str(sheet["F3"].value).split("/")[0]
usernameS2=sheet["B3"].value
en_passwordS2=sheet["D3"].value
passwordS2=sheet["C3"].value

SWInit2=["enable","configure terminal","hostname SW2","crypto key generate rsa","login local","password " + str(passwordS2),"exit","username " + str(usernameS2) + "password " + str(passwordS2)]
with serial.Serial(port="COM4") as console:

    if console.isOpen():
        print("Connected successfully")
        for i in range(len(SWInit)):
         console.write(b"i \n")
         lis.append(i)

         time.sleep(1)
         console.write(b"int vlan 1\n")
         time.sleep(1)
         console.write(b"ip address  " + hostS2 + " 255.255.255.0 \n")
         time.sleep(1)
         console.write(b"no shutdown\n")
         time.sleep(1)
         console.write(b"ip default-gateway" + host2 + "\n")
         time.sleep(1)
         numberofBytes = console.inWaiting()
         data = console.read(numberofBytes)
         print(data.decode())

    else:
        print("You can't connect")
#
work_book3=load_workbook(filename="Switch 2 ports .xlsx")
sheet3=work_book3["Sheet1"]
#
mycon=Netmiko(host=hostS1, username=str(usernameS1), password=str(passwordS1), device_type="cisco_ios", secret= str(en_passwordS1))
mycon.enable()

for x in range(2,11):

    port_number = sheet3["A" + str(x)].value
    port_type = sheet3["B" + str(x)].value
    if(port_type == "Access"):


        commAccess1 = ["vlan " + str(sheet3["C" + str(x)].value),"exit", "interface ethernet 0/" + str(port_number), "switchport mode access" ,"switchport access vlan " + str(sheet3["C" + str(x)].value)]
        # output = mycon.send_config_set(commAccess)
        print(commAccess1)
    else:
        commTrunk1 = ["interface ethernet 0/" + str(port_number), "switchport trunk encapsulation dot1q" , "switchport mode trunk","switchport trunk allowed Vlan " + str(sheet3["C" + str(x)].value)]
        # output = mycon.send_config_set(commTrunk)
        print(commTrunk1)
mycon.disconnect()